"""Exports (Reports phase, Pass 5) — one renderer per format, one source.

The golden rule: exports consume the SAME ``ReportData`` the screen displays.
There is no second calculation path, so the file can never disagree with the
report on screen.

    SalesReport(filters) -> ReportData -> csv | xlsx | pdf
"""
from __future__ import annotations

import io

from ..reports.service import ReportData


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def render_csv(report: ReportData) -> bytes:
    import csv

    buf = io.StringIO()
    w = csv.writer(buf)

    w.writerow(["Co-op", report.title])
    w.writerow(["Period", report.period_label])
    if report.compare != "none":
        w.writerow(["Compared with", report.compare.replace("_", " ")])
    w.writerow(["Generated", report.generated_at])
    f = report.filters
    if f.get("category"):
        w.writerow(["Filter", f"Category: {f['category']}"])
    if f.get("product_id"):
        w.writerow(["Filter", f"Product #{f['product_id']}"])
    if f.get("customer_id"):
        w.writerow(["Filter", f"Customer #{f['customer_id']}"])
    w.writerow([])

    w.writerow(["Key", "Value", "Previous", "Change %"])
    for k in report.kpis:
        w.writerow([k.label, k.value,
                    k.previous if k.previous is not None else "",
                    k.change_percent if k.change_percent is not None else ""])
    w.writerow([])

    for t in report.tables:
        w.writerow([t.title])
        w.writerow(t.columns)
        for row in t.rows:
            w.writerow(row)
        w.writerow([])

    for n in report.notes:
        w.writerow(["Note", n])

    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# XLSX (openpyxl) — Summary sheet + one sheet per table
# ---------------------------------------------------------------------------

def render_xlsx(report: ReportData) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="EEF0FF")

    def _style_header(ws, row_idx: int, n_cols: int) -> None:
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = bold
            cell.fill = head_fill

    def _autosize(ws, max_w: int = 40) -> None:
        for col in ws.columns:
            letter = get_column_letter(col[0].column)
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[letter].width = min(max_w, width + 2)

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws.append([report.title])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append(["Period", report.period_label])
    if report.compare != "none":
        ws.append(["Compared with", report.compare.replace("_", " ")])
    ws.append(["Generated", report.generated_at])
    f = report.filters
    for label, val in (("Category", f.get("category")),
                       ("Product", f.get("product_id")),
                       ("Customer", f.get("customer_id"))):
        if val:
            ws.append([label, val])
    ws.append([])
    ws.append(["KPI", "Value", "Previous", "Change %"])
    _style_header(ws, ws.max_row, 4)
    for k in report.kpis:
        ws.append([k.label, k.value,
                   k.previous if k.previous is not None else None,
                   k.change_percent if k.change_percent is not None else None])
    if report.notes:
        ws.append([])
        for n in report.notes:
            ws.append([n])
    _autosize(ws)

    # --- One sheet per table ---
    used: set[str] = set()
    for t in report.tables:
        base = (t.title[:28] or "Table").strip()
        name = base
        i = 2
        while name in used:
            name = f"{base[:24]} {i}"
            i += 1
        used.add(name)
        tws = wb.create_sheet(title=name)
        tws.append(t.columns)
        _style_header(tws, 1, len(t.columns))
        for row in t.rows:
            tws.append(list(row))
        _autosize(tws)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# PDF (reportlab) — presentation-quality summary
# ---------------------------------------------------------------------------

def render_pdf(report: ReportData) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    page = landscape(A4)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=page, leftMargin=16 * mm, rightMargin=16 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
        title=f"Co-op — {report.title}",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], fontSize=20, leading=24, spaceAfter=2)
    sub_style = ParagraphStyle("s", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#555555"))
    sec_style = ParagraphStyle("sec", parent=styles["Heading2"], fontSize=13, leading=16,
                               spaceBefore=10, spaceAfter=4, textColor=colors.HexColor("#3335b8"))

    elements = []
    elements.append(Paragraph("Co-op", ParagraphStyle("brand", parent=styles["Normal"],
                                                      fontSize=12, textColor=colors.HexColor("#5b5fef"),
                                                      spaceAfter=0)))
    elements.append(Paragraph(report.title, title_style))
    meta = [f"Period: {report.period_label}"]
    if report.compare != "none":
        meta.append(f"Compared with: {report.compare.replace('_', ' ')}")
    f = report.filters
    if f.get("category"):
        meta.append(f"Category: {f['category']}")
    elements.append(Paragraph("    ·    ".join(meta), sub_style))
    elements.append(Paragraph(f"Generated {report.generated_at}", sub_style))
    elements.append(Spacer(1, 8))

    # KPI table (label/value/previous/change) — two columns of KPIs.
    kpi_rows = [["KPI", "Value", "Previous", "Change %"]]
    for k in report.kpis:
        kpi_rows.append([
            k.label,
            _fmt_cell(k.value, k.format),
            _fmt_cell(k.previous, k.format) if k.previous is not None else "—",
            _fmt_change(k.change_percent) if k.change_percent is not None else "—",
        ])
    kpi_table = Table(kpi_rows, colWidths=[60 * mm, 45 * mm, 45 * mm, 40 * mm], hAlign="LEFT")
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF0FF")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#C9CCE8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F8FF")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(Paragraph("Key figures", sec_style))
    elements.append(kpi_table)

    # Simple chart: a bar for the primary series if it's a line/bar chart.
    if report.chart.kind in ("line", "bar") and report.chart.series and report.chart.labels and report.chart.series[0]["data"]:
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.graphics.renderPDF import Drawing, GraphicsFlowable

        chart = VerticalBarChart()
        chart.width, chart.height = 150 * mm, 40 * mm
        data = [max(0, float(v or 0)) for v in report.chart.series[0]["data"][:16]]
        labels = report.chart.labels[:16]
        chart.data = [data]
        chart.categoryAxis.categoryNames = [_short(l) for l in labels]
        chart.categoryAxis.labels.fontSize = 6
        chart.valueAxis.valueMin = 0
        chart.bars[0].fillColor = colors.HexColor("#5b5fef")
        chart.barLabels.fontName = "Helvetica"
        chart.barLabels.fontSize = 5
        drawing = Drawing(chart.width + 10, chart.height + 25)
        drawing.add(chart)
        elements.append(Paragraph(f"{report.chart.series[0]['name']} over time", sec_style))
        elements.append(GraphicsFlowable(drawing))

    # Tables.
    for t in report.tables:
        if not t.rows:
            continue
        data = [t.columns] + [list(r) for r in t.rows[:25]]
        n = len(t.columns)
        table = Table(data, hAlign="LEFT", repeatRows=1)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF0FF")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D5D8EC")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for ci in t.numeric_cols:
            if ci < n:
                style.append(("ALIGN", (ci, 1), (ci, -1), "RIGHT"))
        table.setStyle(TableStyle(style))
        elements.append(Paragraph(t.title, sec_style))
        elements.append(table)
        elements.append(Spacer(1, 4))

    if report.notes:
        elements.append(Paragraph("Notes", sec_style))
        for n in report.notes:
            elements.append(Paragraph(n, sub_style))

    doc.build(elements)
    return buf.getvalue()


def _fmt_cell(v, fmt: str) -> str:
    if v is None:
        return "—"
    try:
        num = float(v)
    except (TypeError, ValueError):
        return str(v)
    if fmt == "money":
        return f"{num:,.0f}" if abs(num - round(num)) < 0.05 else f"{num:,.2f}"
    if fmt == "percent":
        return f"{num:,.1f}%"
    return f"{num:,.0f}" if abs(num - round(num)) < 0.05 else f"{num:,.2f}"


def _fmt_change(v) -> str:
    try:
        return f"{float(v):+.1f}%"
    except (TypeError, ValueError):
        return str(v)


def _short(s: str, n: int = 8) -> str:
    s = str(s)
    return s if len(s) <= n else s[: n - 1] + "…"
