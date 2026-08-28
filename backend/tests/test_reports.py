"""Reports engine + exports — numbers matter, so this is tested hard.

Covers calculation correctness, the single filter contract, comparison
periods, export consistency (screen == file) and the empty-data edge cases.
"""
from __future__ import annotations

import datetime as dt

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from backend.exports import export_report
from backend.models import (
    Base, Business, Customer, Order, OrderItem, OrderStatus, Product, StockMovement,
)
from backend.reports import FilterError, ReportFilters, build_report

TODAY = dt.date.today()


def _d(days_ago: int) -> dt.datetime:
    return dt.datetime.combine(TODAY - dt.timedelta(days=days_ago), dt.time(12))


async def _seed(session_factory):
    """Deterministic fixture:
    - Chair (Furniture, $100, cost $40, stock 3, reorder 5)
    - Desk  (Furniture, $200, cost $90, stock 0, reorder 2)
    - Lamp  (Lighting,  $30,  cost None, stock 20, reorder 5)
    - Alice (2 orders in period), Bob (1 order 40 days ago)
    Returns (business_id, product_ids, customer_ids).
    """
    async with session_factory() as db:
        b = Business(name="Test", owner_id="u1")
        db.add(b)
        await db.flush()
        chair = Product(business_id=b.id, name="Chair", sku="CH1", category="Furniture",
                        unit_price=100, cost_price=40, current_stock=3, reorder_level=5)
        desk = Product(business_id=b.id, name="Desk", sku="DS1", category="Furniture",
                       unit_price=200, cost_price=90, current_stock=0, reorder_level=2)
        lamp = Product(business_id=b.id, name="Lamp", sku="LP1", category="Lighting",
                       unit_price=30, cost_price=None, current_stock=20, reorder_level=5)
        alice = Customer(business_id=b.id, full_name="Alice", email="a@x.com")
        bob = Customer(business_id=b.id, full_name="Bob", email="b@x.com")
        db.add_all([chair, desk, lamp, alice, bob])
        await db.flush()

        def order(cid, pid, qty, price, days_ago, oid_holder):
            o = Order(business_id=b.id, customer_id=cid, status=OrderStatus.delivered,
                      total_amount=qty * price, order_date=_d(days_ago))
            db.add(o)
            oid_holder.append(o)

        oids = []
        order(alice.id, chair.id, 2, 100, 5, oids)
        await db.flush()
        db.add(OrderItem(business_id=b.id, order_id=oids[0].id, product_id=chair.id,
                         quantity=2, unit_price=100, total_price=200))
        order(alice.id, desk.id, 1, 200, 10, oids)
        await db.flush()
        db.add(OrderItem(business_id=b.id, order_id=oids[1].id, product_id=desk.id,
                         quantity=1, unit_price=200, total_price=200))
        order(bob.id, lamp.id, 4, 30, 40, oids)
        await db.flush()
        db.add(OrderItem(business_id=b.id, order_id=oids[2].id, product_id=lamp.id,
                         quantity=4, unit_price=30, total_price=120))
        db.add(StockMovement(business_id=b.id, product_id=chair.id, change=-3, reason="order"))
        await db.commit()
        return b.id, [chair.id, desk.id, lamp.id], [alice.id, bob.id]


@pytest.fixture
async def seeded(session_factory):
    return await _seed(session_factory)


def _kpi(r, key):
    return next(k for k in r.kpis if k.key == key)


# ---------------------------------------------------------------------------
# Calculation correctness
# ---------------------------------------------------------------------------

async def test_sales_kpis(session_factory, seeded):
    bid, pids, cids = seeded
    async with session_factory() as db:
        r = await build_report(db, bid, "sales", ReportFilters.from_query())
    assert _kpi(r, "revenue").value == 400.0   # 200 + 200 (Lamp is 40 days out)
    assert _kpi(r, "orders").value == 2
    assert _kpi(r, "units").value == 3
    assert _kpi(r, "aov").value == 200.0


async def test_sales_comparison_previous_period(session_factory, seeded):
    bid, _, _ = seeded
    async with session_factory() as db:
        r = await build_report(db, bid, "sales", ReportFilters.from_query(compare="previous_period"))
    rev = _kpi(r, "revenue")
    assert rev.value == 400.0
    assert rev.previous == 120.0            # the 40-day-ago Lamp order
    assert rev.change_percent == 233.3      # (400-120)/120


async def test_profit_loss_cogs_and_margin(session_factory, seeded):
    bid, _, _ = seeded
    async with session_factory() as db:
        r = await build_report(db, bid, "profit-loss", ReportFilters.from_query())
    assert _kpi(r, "revenue").value == 400.0
    assert _kpi(r, "cogs").value == 170.0   # Chair 2*40=80 + Desk 1*90=90 (Lamp is 40d out)
    assert _kpi(r, "gross_profit").value == 230.0
    assert _kpi(r, "gross_margin").value == 57.5
    # Honest P&L positioning note is always present (no operating expenses).
    assert any("operating expenses" in n for n in r.notes)


async def test_profit_loss_cost_coverage_note(session_factory, seeded):
    """When a no-cost product is inside the window, COGS coverage is disclosed."""
    bid, _, _ = seeded
    # Widen the window to include the 40-day-ago Lamp order (no cost price).
    f = ReportFilters.from_query(
        from_str=(TODAY - dt.timedelta(days=45)).isoformat(),
        to_str=TODAY.isoformat(),
    )
    async with session_factory() as db:
        r = await build_report(db, bid, "profit-loss", f)
    assert _kpi(r, "revenue").value == 520.0  # 400 + 120
    assert _kpi(r, "cogs").value == 170.0     # Lamp has no cost -> excluded
    assert any("Cost data covers" in n for n in r.notes)


async def test_inventory_value_and_risk(session_factory, seeded):
    bid, _, _ = seeded
    async with session_factory() as db:
        r = await build_report(db, bid, "inventory", ReportFilters.from_query())
    assert _kpi(r, "value").value == 720.0  # Chair 3*40=120 + Lamp 20*30=600
    assert _kpi(r, "units").value == 23
    assert _kpi(r, "low").value == 1        # Chair (3 <= 5)
    assert _kpi(r, "out").value == 1        # Desk (0)


async def test_customers_repeat_and_revenue(session_factory, seeded):
    bid, _, cids = seeded
    async with session_factory() as db:
        r = await build_report(db, bid, "customers", ReportFilters.from_query())
    assert _kpi(r, "total").value == 2
    assert _kpi(r, "repeat").value == 1     # Alice has 2 orders in period
    assert _kpi(r, "rev_per_cust").value == 400.0  # only Alice sold in period


# ---------------------------------------------------------------------------
# Filter contract (one engine, one contract)
# ---------------------------------------------------------------------------

async def test_filter_category_is_line_level(session_factory, seeded):
    bid, _, _ = seeded
    async with session_factory() as db:
        furn = await build_report(db, bid, "sales", ReportFilters.from_query(category="Furniture"))
        light = await build_report(db, bid, "sales", ReportFilters.from_query(category="Lighting"))
    assert _kpi(furn, "revenue").value == 400.0
    assert _kpi(light, "revenue").value == 0  # Lamp order is 40 days out of the 30d window


async def test_filter_customer(session_factory, seeded):
    bid, _, cids = seeded
    async with session_factory() as db:
        alice = await build_report(db, bid, "sales", ReportFilters.from_query(customer_id=cids[0]))
    assert _kpi(alice, "revenue").value == 400.0
    assert _kpi(alice, "orders").value == 2


async def test_filter_product(session_factory, seeded):
    bid, pids, _ = seeded
    async with session_factory() as db:
        chair = await build_report(db, bid, "sales", ReportFilters.from_query(product_id=pids[0]))
    assert _kpi(chair, "revenue").value == 200.0


async def test_invalid_date_range_rejected():
    with pytest.raises(FilterError):
        ReportFilters.from_query(from_str="2026-08-28", to_str="2026-01-01")
    with pytest.raises(FilterError):
        ReportFilters.from_query(from_str="not-a-date")
    with pytest.raises(FilterError):
        ReportFilters.from_query(compare="nonsense")


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------

async def test_empty_business_all_zero(session_factory):
    async with session_factory() as db:
        b = Business(name="Empty", owner_id="u2")
        db.add(b)
        await db.commit()
        bid = b.id
    f = ReportFilters.from_query()
    async with session_factory() as db:
        for key in ("sales", "profit-loss", "inventory", "customers"):
            r = await build_report(db, bid, key, f)
            for k in r.kpis:
                assert k.value in (0, None)
            assert r.to_dict()["tables"]  # still well-formed


# ---------------------------------------------------------------------------
# Export consistency — the file must carry the same numbers as the screen
# ---------------------------------------------------------------------------

async def test_export_csv_matches_report(session_factory, seeded):
    bid, _, _ = seeded
    async with session_factory() as db:
        r = await build_report(db, bid, "sales", ReportFilters.from_query(compare="previous_period"))
    content, fname, ct = export_report(r, "csv")
    text = content.decode("utf-8")
    assert ct.startswith("text/csv")
    assert fname.endswith(".csv")
    # KPI values present, and the comparison change present.
    assert "400" in text and "233.3" in text
    assert "Revenue" in text and "Top products" in text


async def test_export_xlsx_summary_matches_report(session_factory, seeded):
    bid, _, _ = seeded
    async with session_factory() as db:
        r = await build_report(db, bid, "sales", ReportFilters.from_query(compare="previous_period"))
    content, fname, ct = export_report(r, "xlsx")
    wb = load_workbook(__import__("io").BytesIO(content))
    summary = wb["Summary"]
    vals = {summary.cell(row=i, column=1).value: summary.cell(row=i, column=2).value
            for i in range(1, summary.max_row + 1) if summary.cell(row=i, column=1).value}
    assert vals.get("Revenue") == 400.0
    assert vals.get("Orders") == 2
    # One sheet per report table.
    assert "Top products" in wb.sheetnames
    assert "Sales by category" in wb.sheetnames


async def test_export_pdf_is_valid(session_factory, seeded):
    bid, _, _ = seeded
    async with session_factory() as db:
        r = await build_report(db, bid, "profit-loss", ReportFilters.from_query())
    content, fname, ct = export_report(r, "pdf")
    assert ct == "application/pdf"
    assert content[:4] == b"%PDF"
    assert len(content) > 2000  # real content, not a stub


def test_unknown_format_rejected():
    from backend.exports import ExportError
    from backend.reports.service import ReportData
    r = ReportData(key="sales", title="Sales", period_label="x", compare="none",
                   filters={}, kpis=[], chart=None, tables=[])
    with pytest.raises(ExportError):
        export_report(r, "docx")


# ---------------------------------------------------------------------------
# Filter previous_range logic
# ---------------------------------------------------------------------------

def test_previous_period_range():
    f = ReportFilters.from_query(from_str="2026-08-01", to_str="2026-08-30", compare="previous_period")
    prev = f.previous_range()
    # 30-day window ending the day before `from`, same length.
    assert prev == (dt.date(2026, 7, 2), dt.date(2026, 7, 31))


def test_previous_month_range():
    f = ReportFilters.from_query(from_str="2026-08-01", to_str="2026-08-30", compare="previous_month")
    prev = f.previous_range()
    assert prev == (dt.date(2026, 7, 1), dt.date(2026, 7, 31))


def test_none_compare_has_no_previous():
    f = ReportFilters.from_query(from_str="2026-08-01", to_str="2026-08-30")
    assert f.previous_range() is None
